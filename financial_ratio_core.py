from __future__ import annotations

import io
import json
import re
import tempfile
import textwrap
from pathlib import Path
from typing import Iterable, Sequence

import httpx
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import wrds
from matplotlib import font_manager
from matplotlib.ticker import PercentFormatter
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

try:
    from wrds.sql import WRDS_CONNECT_ARGS, WRDS_POSTGRES_DB, WRDS_POSTGRES_HOST, WRDS_POSTGRES_PORT
except Exception:
    WRDS_POSTGRES_HOST = "wrds-pgdata.wharton.upenn.edu"
    WRDS_POSTGRES_PORT = 9737
    WRDS_POSTGRES_DB = "wrds"
    WRDS_CONNECT_ARGS = {"sslmode": "require"}

TAX_RATE = 0.21
WRDS_CONNECTION_TIMEOUT_SECONDS = 20
CHART_FONT_CANDIDATES = [
    "Times New Roman",
    "Times",
    "Nimbus Roman",
    "Nimbus Roman No9 L",
    "TeX Gyre Termes",
    "Liberation Serif",
    "STIXGeneral",
    "DejaVu Serif",
]

METRIC_GROUPS = {
    "Performance Measures": [
        "MARKET_VALUE_ADDED",
        "MARKET_TO_BOOK_RATIO",
    ],
    "Profitability Measures": [
        "ROA",
        "ROC",
        "ROE",
        "EVA",
    ],
    "Efficiency Measures": [
        "ASSET_TURNOVER",
        "RECEIVABLES_TURNOVER",
        "AVERAGE_COLLECTION_PERIOD_DAYS",
        "INVENTORY_TURNOVER",
        "DAYS_IN_INVENTORY",
        "PROFIT_MARGIN",
        "OPERATING_PROFIT_MARGIN",
    ],
    "Leverage Measures": [
        "LONG_TERM_DEBT_RATIO",
        "LONG_TERM_DEBT_EQUITY_RATIO",
        "TOTAL_DEBT_RATIO",
        "TIMES_INTEREST_EARNED",
        "CASH_COVERAGE_RATIO",
    ],
    "Liquidity Measures": [
        "NET_WORKING_CAPITAL_TO_ASSETS",
        "CURRENT_RATIO",
        "QUICK_RATIO",
        "CASH_RATIO",
    ],
}

ALL_METRICS = [metric for group in METRIC_GROUPS.values() for metric in group]

DEFAULT_PERCENT_METRICS = [
    "ROA",
    "ROC",
    "ROE",
    "PROFIT_MARGIN",
    "OPERATING_PROFIT_MARGIN",
    "LONG_TERM_DEBT_RATIO",
    "TOTAL_DEBT_RATIO",
    "NET_WORKING_CAPITAL_TO_ASSETS",
]

METRIC_LABELS = {
    "MARKET_VALUE_ADDED": "Market Value Added",
    "MARKET_TO_BOOK_RATIO": "Market-to-book Ratio",
    "ROA": "ROA",
    "ROC": "ROC",
    "ROE": "ROE",
    "EVA": "EVA",
    "ASSET_TURNOVER": "Asset Turnover",
    "RECEIVABLES_TURNOVER": "Receivables Turnover",
    "AVERAGE_COLLECTION_PERIOD_DAYS": "Average Collection Period (days)",
    "INVENTORY_TURNOVER": "Inventory Turnover",
    "DAYS_IN_INVENTORY": "Days in Inventory",
    "PROFIT_MARGIN": "Profit Margin",
    "OPERATING_PROFIT_MARGIN": "Operating Profit Margin",
    "LONG_TERM_DEBT_RATIO": "Long-term Debt Ratio",
    "LONG_TERM_DEBT_EQUITY_RATIO": "Long-term Debt-equity Ratio",
    "TOTAL_DEBT_RATIO": "Total Debt Ratio",
    "TIMES_INTEREST_EARNED": "Times Interest Earned",
    "CASH_COVERAGE_RATIO": "Cash Coverage Ratio",
    "NET_WORKING_CAPITAL_TO_ASSETS": "Net Working Capital to Assets",
    "CURRENT_RATIO": "Current Ratio",
    "QUICK_RATIO": "Quick Ratio",
    "CASH_RATIO": "Cash Ratio",
}

METRIC_REFERENCE = {
    "MARKET_VALUE_ADDED": {
        "group": "Performance Measures",
        "formula": "Market value of equity - book equity",
        "meaning": "Shows how much value the market places on the company above its recorded book equity.",
        "notes": "",
    },
    "MARKET_TO_BOOK_RATIO": {
        "group": "Performance Measures",
        "formula": "Market value of equity / book equity",
        "meaning": "Shows how strongly the market values the company relative to its accounting equity base.",
        "notes": "",
    },
    "ROA": {
        "group": "Profitability Measures",
        "formula": "After-tax operating income / total assets",
        "meaning": "Shows how efficiently the firm's assets generate profit.",
        "notes": "Fallback: if after-tax operating income cannot be formed, use net income / total assets.",
    },
    "ROC": {
        "group": "Profitability Measures",
        "formula": "After-tax operating income / capital, where capital = long-term debt + equity",
        "meaning": "Shows the return earned on the long-term capital invested in the business.",
        "notes": "If long-term debt is missing, it is treated as 0 when building capital.",
    },
    "ROE": {
        "group": "Profitability Measures",
        "formula": "Net income / equity",
        "meaning": "Shows the return generated for shareholders from the equity base.",
        "notes": "",
    },
    "EVA": {
        "group": "Profitability Measures",
        "formula": "After-tax operating income - cost of capital x capital",
        "meaning": "Shows whether operating profit exceeds the required return demanded by capital providers.",
        "notes": "Requires the user to provide a cost of capital.",
    },
    "ASSET_TURNOVER": {
        "group": "Efficiency Measures",
        "formula": "Sales / total assets",
        "meaning": "Shows how efficiently total assets are used to generate revenue.",
        "notes": "",
    },
    "RECEIVABLES_TURNOVER": {
        "group": "Efficiency Measures",
        "formula": "Sales / receivables",
        "meaning": "Shows how quickly receivables are converted into revenue over the year.",
        "notes": "",
    },
    "AVERAGE_COLLECTION_PERIOD_DAYS": {
        "group": "Efficiency Measures",
        "formula": "Receivables / (sales / 365)",
        "meaning": "Estimates the average number of days customers take to pay.",
        "notes": "",
    },
    "INVENTORY_TURNOVER": {
        "group": "Efficiency Measures",
        "formula": "Cost of goods sold / inventory",
        "meaning": "Shows how quickly inventory is sold and replaced.",
        "notes": "",
    },
    "DAYS_IN_INVENTORY": {
        "group": "Efficiency Measures",
        "formula": "Inventory / (cost of goods sold / 365)",
        "meaning": "Estimates the average number of days inventory remains on hand.",
        "notes": "",
    },
    "PROFIT_MARGIN": {
        "group": "Efficiency Measures",
        "formula": "Net income / sales",
        "meaning": "Shows how much of each sales dollar is kept as net profit.",
        "notes": "",
    },
    "OPERATING_PROFIT_MARGIN": {
        "group": "Efficiency Measures",
        "formula": "After-tax operating income / sales",
        "meaning": "Shows how much operating profit the company keeps from revenue after tax effects.",
        "notes": "",
    },
    "LONG_TERM_DEBT_RATIO": {
        "group": "Leverage Measures",
        "formula": "Long-term debt / capital",
        "meaning": "Shows the share of long-term capital financed by long-term debt.",
        "notes": "If long-term debt is missing, it is treated as 0.",
    },
    "LONG_TERM_DEBT_EQUITY_RATIO": {
        "group": "Leverage Measures",
        "formula": "Long-term debt / equity",
        "meaning": "Compares long-term debt financing to shareholders' equity.",
        "notes": "If long-term debt is missing, it is treated as 0.",
    },
    "TOTAL_DEBT_RATIO": {
        "group": "Leverage Measures",
        "formula": "Total liabilities / total assets",
        "meaning": "Shows the proportion of assets financed by liabilities.",
        "notes": "",
    },
    "TIMES_INTEREST_EARNED": {
        "group": "Leverage Measures",
        "formula": "EBIT / interest expense",
        "meaning": "Shows how many times operating earnings can cover interest payments.",
        "notes": "Left blank when interest expense is missing or zero to avoid a misleading ratio.",
    },
    "CASH_COVERAGE_RATIO": {
        "group": "Leverage Measures",
        "formula": "(EBIT + depreciation) / interest expense",
        "meaning": "Shows interest-paying ability after adding back depreciation as a non-cash charge.",
        "notes": "Depreciation defaults to 0 if missing. Left blank when interest expense is missing or zero.",
    },
    "NET_WORKING_CAPITAL_TO_ASSETS": {
        "group": "Liquidity Measures",
        "formula": "(Current assets - current liabilities) / total assets",
        "meaning": "Shows the portion of assets financed by net short-term operating liquidity.",
        "notes": "",
    },
    "CURRENT_RATIO": {
        "group": "Liquidity Measures",
        "formula": "Current assets / current liabilities",
        "meaning": "Shows the firm's ability to cover short-term obligations with short-term assets.",
        "notes": "",
    },
    "QUICK_RATIO": {
        "group": "Liquidity Measures",
        "formula": "(Cash + short-term investments + receivables) / current liabilities",
        "meaning": "Shows short-term liquidity without relying on inventory.",
        "notes": "",
    },
    "CASH_RATIO": {
        "group": "Liquidity Measures",
        "formula": "(Cash + short-term investments) / current liabilities",
        "meaning": "Shows the most conservative short-term liquidity position using only near-cash assets.",
        "notes": "",
    },
}


def get_metric_label(metric: str) -> str:
    metric = metric.upper()
    return METRIC_LABELS.get(metric, metric.replace("_", " ").title())


def clean_metrics(metric_list: Sequence[str] | str) -> list[str]:
    if metric_list is None:
        raise ValueError("Please choose at least one financial metric.")

    if isinstance(metric_list, str):
        metric_list = [metric_list]

    metric_list = [str(metric).upper().strip() for metric in metric_list if str(metric).strip()]
    if not metric_list:
        raise ValueError("Please choose at least one financial metric.")

    if "ALL" in metric_list:
        return list(ALL_METRICS)

    invalid = [metric for metric in metric_list if metric not in ALL_METRICS]
    if invalid:
        raise ValueError(f"Unsupported metric(s): {invalid}")

    return metric_list


def parse_tickers(tickers: Sequence[str] | str) -> list[str]:
    if isinstance(tickers, str):
        pieces = re.split(r"[\s,;\n]+", tickers.strip())
    else:
        pieces = [str(ticker).strip() for ticker in tickers]

    cleaned = []
    for ticker in pieces:
        if ticker:
            cleaned.append(ticker.upper())

    cleaned = list(dict.fromkeys(cleaned))
    if not cleaned:
        raise ValueError("Please enter at least one ticker code.")

    return cleaned


def get_pgpass_path() -> Path:
    import os

    if os.name == "nt":
        appdata = Path(os.getenv("APPDATA", str(Path.home())))
        return appdata / "postgresql" / "pgpass.conf"
    return Path.home() / ".pgpass"


def has_pgpass_entry(
    username: str,
    wrds_hostname: str = WRDS_POSTGRES_HOST,
    wrds_port: int = WRDS_POSTGRES_PORT,
    wrds_dbname: str = WRDS_POSTGRES_DB,
) -> bool:
    pgpass_path = get_pgpass_path()
    if not pgpass_path.exists():
        return False

    for raw_line in pgpass_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        fields = line.replace(r"\:", "##COLON##").split(":")
        if len(fields) < 5:
            continue

        if (
            fields[0] == wrds_hostname
            and fields[1] == str(wrds_port)
            and fields[2] == wrds_dbname
            and fields[3] == username
        ):
            return True

    return False


def create_wrds_pgpass_entry(
    username: str,
    password: str,
    wrds_hostname: str = WRDS_POSTGRES_HOST,
    wrds_port: int = WRDS_POSTGRES_PORT,
    wrds_dbname: str = WRDS_POSTGRES_DB,
) -> str:
    if not username or not password:
        raise ValueError("A WRDS username and password are required to create pgpass.conf.")

    db = build_wrds_connection(
        username=username,
        password=password,
        wrds_hostname=wrds_hostname,
        wrds_port=wrds_port,
        wrds_dbname=wrds_dbname,
    )
    try:
        db.create_pgpass_file()
        return str(get_pgpass_path())
    finally:
        try:
            db.close()
        except Exception:
            pass


def build_wrds_connect_args(timeout_seconds: int = WRDS_CONNECTION_TIMEOUT_SECONDS) -> dict:
    connect_args = dict(WRDS_CONNECT_ARGS) if isinstance(WRDS_CONNECT_ARGS, dict) else {"sslmode": "require"}
    connect_args.setdefault("sslmode", "require")
    connect_args["connect_timeout"] = int(timeout_seconds)
    return connect_args


def build_wrds_error_message(exc: Exception, timeout_seconds: int = WRDS_CONNECTION_TIMEOUT_SECONDS) -> str:
    message = str(exc).strip() or exc.__class__.__name__
    lowered = message.lower()

    authentication_markers = [
        "password authentication failed",
        "authentication failed",
        "invalid password",
        "fe_sendauth",
        "28p01",
    ]
    timeout_markers = [
        "timeout expired",
        "timed out",
        "timeout",
    ]

    if any(marker in lowered for marker in authentication_markers):
        return (
            "WRDS login failed. The username or password appears to be incorrect. "
            "Please check the password and try again."
        )

    if any(marker in lowered for marker in timeout_markers):
        return (
            f"WRDS connection timed out after about {timeout_seconds} seconds. "
            "If WRDS prompted for Duo 2FA, please finish the Duo approval first. "
            "Otherwise, check that the username, password, and network connection are correct."
        )

    if "could not connect to server" in lowered:
        return (
            "WRDS could not be reached from the app. Please check your network connection, "
            "then try again."
        )

    return f"WRDS connection failed: {message}"


def build_wrds_connection(
    username: str,
    password: str | None = None,
    wrds_hostname: str = WRDS_POSTGRES_HOST,
    wrds_port: int = WRDS_POSTGRES_PORT,
    wrds_dbname: str = WRDS_POSTGRES_DB,
):
    if not username:
        raise ValueError("Please provide your WRDS username.")

    if not password and not has_pgpass_entry(
        username=username,
        wrds_hostname=wrds_hostname,
        wrds_port=wrds_port,
        wrds_dbname=wrds_dbname,
    ):
        raise ValueError(
            "Streamlit cannot open WRDS's interactive password prompt. "
            "Please enter your password or create pgpass.conf first."
        )

    connection_kwargs = {
        "wrds_username": username,
        "wrds_hostname": wrds_hostname,
        "wrds_port": wrds_port,
        "wrds_dbname": wrds_dbname,
        "wrds_connect_args": build_wrds_connect_args(),
    }
    if password:
        connection_kwargs["wrds_password"] = password

    try:
        db = wrds.Connection(autoconnect=False, **connection_kwargs)
        db._Connection__make_sa_engine_conn(raise_err=True)
        db.load_library_list()
        return db
    except Exception as exc:
        raise RuntimeError(build_wrds_error_message(exc)) from exc


def divide(a: pd.Series, b: pd.Series) -> pd.Series:
    return a.div(b).where((b != 0) & b.notna())


def fetch_funda_data(db, where_condition: str, order_by: str = "tic, fyear") -> pd.DataFrame:
    sql = f"""
        SELECT
            conm,
            tic,
            fyear,
            sich,
            sale,
            cogs,
            ni,
            ebit,
            xint,
            dp,
            at,
            lt,
            act,
            lct,
            che,
            ivst,
            rect,
            invt,
            dltt,
            teq,
            seq,
            ceq,
            prcc_f,
            csho
        FROM comp.funda
        WHERE {where_condition}
          AND indfmt = 'INDL'
          AND datafmt = 'STD'
          AND popsrc = 'D'
          AND consol = 'C'
        ORDER BY {order_by}
    """
    return db.raw_sql(sql)


def calculate_metrics(
    df: pd.DataFrame,
    metric_list: Sequence[str],
    cost_of_capital: float | None = None,
) -> pd.DataFrame:
    df = df.copy()

    if df.empty:
        return df

    metric_list = clean_metrics(metric_list)

    if "EVA" in metric_list and cost_of_capital is None:
        raise ValueError("If you choose EVA, please provide a cost of capital such as 0.10.")

    numeric_cols = [
        "sale",
        "cogs",
        "ni",
        "ebit",
        "xint",
        "dp",
        "at",
        "lt",
        "act",
        "lct",
        "che",
        "ivst",
        "rect",
        "invt",
        "dltt",
        "teq",
        "seq",
        "ceq",
        "prcc_f",
        "csho",
    ]
    for column in numeric_cols:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    # Fallback rules are designed to preserve sensible outputs when a component is
    # missing, without using unrelated variables. For example, ROA can fall back
    # to net income / assets, but not to revenue / assets.
    df["equity"] = df["teq"].fillna(df["seq"]).fillna(df["ceq"])
    df["market_value_equity"] = df["prcc_f"] * df["csho"]
    df["interest_expense_for_operations"] = df["xint"].fillna(0)
    df["depreciation_for_coverage"] = df["dp"].fillna(0)
    df["long_term_debt"] = df["dltt"].fillna(0)

    primary_after_tax_operating_income = df["ni"] + (1 - TAX_RATE) * df["interest_expense_for_operations"]
    nopat_from_ebit = df["ebit"] * (1 - TAX_RATE)
    df["after_tax_operating_income"] = primary_after_tax_operating_income.fillna(nopat_from_ebit).fillna(df["ni"])

    df["capital"] = df["long_term_debt"] + df["equity"]
    df["net_working_capital"] = df["act"] - df["lct"]

    df["MARKET_VALUE_ADDED"] = df["market_value_equity"] - df["equity"]
    df["MARKET_TO_BOOK_RATIO"] = divide(df["market_value_equity"], df["equity"])

    df["ROA"] = divide(df["after_tax_operating_income"], df["at"]).fillna(divide(df["ni"], df["at"]))
    df["ROC"] = divide(df["after_tax_operating_income"], df["capital"])
    df["ROE"] = divide(df["ni"], df["equity"])

    if cost_of_capital is not None:
        df["EVA"] = df["after_tax_operating_income"] - cost_of_capital * df["capital"]

    df["ASSET_TURNOVER"] = divide(df["sale"], df["at"])
    df["RECEIVABLES_TURNOVER"] = divide(df["sale"], df["rect"])
    df["AVERAGE_COLLECTION_PERIOD_DAYS"] = divide(df["rect"], df["sale"] / 365)

    df["INVENTORY_TURNOVER"] = divide(df["cogs"], df["invt"])
    df["DAYS_IN_INVENTORY"] = divide(df["invt"], df["cogs"] / 365)

    df["PROFIT_MARGIN"] = divide(df["ni"], df["sale"])
    df["OPERATING_PROFIT_MARGIN"] = divide(df["after_tax_operating_income"], df["sale"])

    df["LONG_TERM_DEBT_RATIO"] = divide(df["long_term_debt"], df["capital"])
    df["LONG_TERM_DEBT_EQUITY_RATIO"] = divide(df["long_term_debt"], df["equity"])
    df["TOTAL_DEBT_RATIO"] = divide(df["lt"], df["at"])

    df["TIMES_INTEREST_EARNED"] = divide(df["ebit"], df["xint"])
    df["CASH_COVERAGE_RATIO"] = divide(df["ebit"] + df["depreciation_for_coverage"], df["xint"])

    df["NET_WORKING_CAPITAL_TO_ASSETS"] = divide(df["net_working_capital"], df["at"])
    df["CURRENT_RATIO"] = divide(df["act"], df["lct"])
    df["QUICK_RATIO"] = divide(df["che"] + df["ivst"].fillna(0) + df["rect"], df["lct"])
    df["CASH_RATIO"] = divide(df["che"] + df["ivst"].fillna(0), df["lct"])

    return df[["conm", "tic", "fyear"] + metric_list]


def trim_sic_outliers(
    metric_df: pd.DataFrame,
    metrics: Sequence[str],
    lower_q: float = 0.10,
    upper_q: float = 0.90,
    min_obs: int = 10,
) -> pd.DataFrame:
    df = metric_df.copy()

    for metric in metrics:
        cleaned_parts = []
        for _, group in df.groupby("fyear"):
            group_copy = group.copy()
            valid = group_copy[metric].dropna()

            if len(valid) < min_obs:
                cleaned_parts.append(group_copy[[metric]])
                continue

            low = valid.quantile(lower_q)
            high = valid.quantile(upper_q)
            group_copy[metric] = group_copy[metric].where(
                (group_copy[metric] >= low) & (group_copy[metric] <= high)
            )
            cleaned_parts.append(group_copy[[metric]])

        cleaned_metric = pd.concat(cleaned_parts).sort_index()
        df[metric] = cleaned_metric[metric]

    return df


def build_financial_ratio_data(
    username: str,
    tickers: Sequence[str] | str,
    start_year: int,
    end_year: int,
    metrics: Sequence[str] | str,
    sic_code: str | int | None = None,
    cost_of_capital: float | None = None,
    password: str | None = None,
    create_pgpass: bool = False,
    wrds_hostname: str = WRDS_POSTGRES_HOST,
    wrds_port: int = WRDS_POSTGRES_PORT,
    wrds_dbname: str = WRDS_POSTGRES_DB,
) -> dict:
    metrics = clean_metrics(metrics)
    tickers = parse_tickers(tickers)

    if start_year > end_year:
        raise ValueError("The start year must be earlier than or equal to the end year.")

    if "EVA" in metrics and cost_of_capital is None:
        raise ValueError("EVA requires a cost of capital.")

    if create_pgpass:
        create_wrds_pgpass_entry(
            username=username,
            password=password or "",
            wrds_hostname=wrds_hostname,
            wrds_port=wrds_port,
            wrds_dbname=wrds_dbname,
        )

    safe_tickers = [ticker.replace("'", "''") for ticker in tickers]
    ticker_sql = ", ".join(f"'{ticker}'" for ticker in safe_tickers)
    company_raw = pd.DataFrame()
    company_df = pd.DataFrame()
    sic_raw = pd.DataFrame()
    sic_avg_df = pd.DataFrame()

    db = build_wrds_connection(
        username=username,
        password=password,
        wrds_hostname=wrds_hostname,
        wrds_port=wrds_port,
        wrds_dbname=wrds_dbname,
    )

    try:
        company_where = f"""
            tic IN ({ticker_sql})
            AND fyear >= {int(start_year)}
            AND fyear <= {int(end_year)}
        """
        company_raw = fetch_funda_data(db, company_where, order_by="tic, fyear")
        company_df = calculate_metrics(company_raw, metrics, cost_of_capital)

        if sic_code is not None and str(sic_code).strip():
            sic_code_str = str(sic_code).strip()
            if sic_code_str.isdigit():
                sic_condition = f"sich = {sic_code_str}"
            else:
                safe_sic = sic_code_str.replace("'", "''")
                sic_condition = f"sich = '{safe_sic}'"

            sic_where = f"""
                {sic_condition}
                AND fyear >= {int(start_year)}
                AND fyear <= {int(end_year)}
            """
            sic_raw = fetch_funda_data(db, sic_where, order_by="fyear")
            sic_metric_df = calculate_metrics(sic_raw, metrics, cost_of_capital)

            if not sic_metric_df.empty:
                sic_metric_df = trim_sic_outliers(
                    metric_df=sic_metric_df,
                    metrics=metrics,
                    lower_q=0.10,
                    upper_q=0.90,
                    min_obs=10,
                )
                sic_avg_df = sic_metric_df.groupby("fyear")[metrics].mean().reset_index()
                sic_avg_df["conm"] = f"SIC {sic_code_str} AVERAGE"
                sic_avg_df["tic"] = ""
                sic_avg_df = sic_avg_df[["conm", "tic", "fyear"] + metrics]
                long_table = pd.concat([company_df, sic_avg_df], ignore_index=True)
            else:
                long_table = company_df.copy()
        else:
            long_table = company_df.copy()

        if long_table.empty:
            return {
                "metrics": metrics,
                "requested_tickers": tickers,
                "start_year": start_year,
                "end_year": end_year,
                "sic_code": sic_code,
                "company_raw": company_raw,
                "company_metrics": company_df,
                "industry_raw": sic_raw,
                "industry_metrics": sic_avg_df,
                "long_table": long_table,
                "long_table_display": long_table,
                "pivot_tables": {},
            }

        long_table = long_table.copy()
        long_table["row_label"] = np.where(
            long_table["tic"] == "",
            long_table["conm"],
            long_table["conm"] + " (" + long_table["tic"] + ")",
        )

        row_order = long_table["row_label"].drop_duplicates().tolist()
        pivot_tables = {}
        for metric in metrics:
            pivot_df = long_table.pivot(index="row_label", columns="fyear", values=metric)
            pivot_df = pivot_df.reindex(row_order)
            pivot_df = pivot_df.reindex(sorted(pivot_df.columns), axis=1)
            pivot_df.index.name = None
            pivot_df.columns.name = None
            pivot_tables[metric] = pivot_df

        long_table_no_label = long_table.drop(columns=["row_label"])
        long_table_display = long_table_no_label.rename(
            columns={metric: get_metric_label(metric) for metric in metrics}
        )

        return {
            "metrics": metrics,
            "requested_tickers": tickers,
            "start_year": start_year,
            "end_year": end_year,
            "sic_code": sic_code,
            "company_raw": company_raw,
            "company_metrics": company_df,
            "industry_raw": sic_raw,
            "industry_metrics": sic_avg_df,
            "long_table": long_table_no_label,
            "long_table_display": long_table_display,
            "pivot_tables": pivot_tables,
        }
    finally:
        db.close()


def join_company_names(names: Sequence[str]) -> str:
    names = [name for name in names if name]
    if not names:
        return "Selected Companies"
    if len(names) == 1:
        return names[0]
    if len(names) == 2:
        return names[0] + " and " + names[1]
    return ", ".join(names[:-1]) + " and " + names[-1]


def build_metric_texts(
    report: dict,
    metric_name: str,
    source_text: str = "Compustat - North America",
) -> tuple[str, str, str]:
    long_df = report["long_table"].copy()
    company_names = long_df.loc[long_df["tic"] != "", "conm"].drop_duplicates().tolist()
    company_text = join_company_names(company_names)

    years = sorted(report["pivot_tables"][metric_name].columns.tolist())
    first_year = years[0]
    last_year = years[-1]
    metric_label = get_metric_label(metric_name)

    title = f"{metric_label} for {company_text} in {first_year}-{last_year}"
    notes = (
        f"Note(s): This table presents the {metric_label} for {company_text} "
        f"in the years of {first_year} to {last_year}."
    )
    source = f"Source(s): {source_text}"
    return title, notes, source


def format_pivot_table(
    pivot_df: pd.DataFrame,
    metric_name: str,
    percent_metrics: Sequence[str] | None = None,
    decimals: int = 2,
) -> pd.DataFrame:
    percent_metrics = [metric.upper() for metric in (percent_metrics or DEFAULT_PERCENT_METRICS)]
    metric_name = metric_name.upper()

    if metric_name in percent_metrics:
        return pivot_df.map(lambda value: f"{value * 100:.{decimals}f}%" if pd.notna(value) else "")

    return pivot_df.map(lambda value: f"{value:,.{decimals}f}" if pd.notna(value) else "")


def build_presentable_metric_table(
    display_df: pd.DataFrame,
    title: str,
    notes: str,
    source: str,
) -> pd.DataFrame:
    table_df = display_df.reset_index().rename(columns={"index": "Item"})
    original_columns = table_df.columns.tolist()
    label_column = original_columns[0]

    header_row = {col: col for col in original_columns}
    title_row = {col: "" for col in original_columns}
    note_row = {col: "" for col in original_columns}
    source_row = {col: "" for col in original_columns}

    header_row[label_column] = ""
    title_row[label_column] = title
    note_row[label_column] = notes
    source_row[label_column] = source

    final_df = pd.concat(
        [
            pd.DataFrame([title_row]),
            pd.DataFrame([header_row]),
            table_df,
            pd.DataFrame([note_row]),
            pd.DataFrame([source_row]),
        ],
        ignore_index=True,
    )
    final_df.columns = [""] * len(final_df.columns)
    return final_df


def _escape_markdown_table_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def dataframe_to_markdown(display_df: pd.DataFrame) -> str:
    try:
        return display_df.to_markdown()
    except Exception as exc:
        if "tabulate" not in str(exc).lower():
            raise

        headers = [""] + [_escape_markdown_table_cell(column) for column in display_df.columns.tolist()]
        separator = ["---"] * len(headers)
        lines = [
            "| " + " | ".join(headers) + " |",
            "| " + " | ".join(separator) + " |",
        ]

        for index_value, row in display_df.iterrows():
            row_cells = [_escape_markdown_table_cell(index_value)] + [
                _escape_markdown_table_cell(value) for value in row.tolist()
            ]
            lines.append("| " + " | ".join(row_cells) + " |")

        return "\n".join(lines)


def build_table_package(
    report: dict,
    decimals: int = 2,
    percent_metrics: Sequence[str] | None = None,
    source_text: str = "Compustat - North America",
) -> dict:
    display_tables = {}
    markdown_tables = {}
    presentable_tables = {}
    titles = {}
    notes = {}
    sources = {}

    for metric, pivot_df in report["pivot_tables"].items():
        display_df = format_pivot_table(
            pivot_df=pivot_df,
            metric_name=metric,
            percent_metrics=percent_metrics,
            decimals=decimals,
        )
        title, note_text, source_line = build_metric_texts(
            report=report,
            metric_name=metric,
            source_text=source_text,
        )
        final_table = build_presentable_metric_table(
            display_df=display_df,
            title=title,
            notes=note_text,
            source=source_line,
        )

        display_tables[metric] = display_df
        markdown_tables[metric] = dataframe_to_markdown(display_df)
        presentable_tables[metric] = final_table
        titles[metric] = title
        notes[metric] = note_text
        sources[metric] = source_line

    report["display_tables"] = display_tables
    report["markdown_tables"] = markdown_tables
    report["presentable_tables"] = presentable_tables
    report["titles"] = titles
    report["notes"] = notes
    report["sources"] = sources
    return report


def export_presentable_table_to_excel(final_df: pd.DataFrame, filename: str, sheet_name: str = "Table") -> None:
    final_df.to_excel(filename, index=False, header=False, sheet_name=sheet_name)

    workbook = load_workbook(filename)
    worksheet = workbook[sheet_name]

    max_col = worksheet.max_column
    max_row = worksheet.max_row

    title_row = 1
    header_row = 2
    notes_row = max_row - 1
    source_row = max_row
    body_start_row = 3
    body_end_row = max_row - 2

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    black_medium = Side(style="medium", color="000000")
    black_thin = Side(style="thin", color="000000")

    title_border = Border(bottom=black_medium)
    header_border = Border(bottom=black_thin)
    item_header_border = Border()
    notes_border = Border(top=black_medium)

    font_12 = Font(name="Times New Roman", size=12, bold=False)
    font_12_bold = Font(name="Times New Roman", size=12, bold=True)
    font_10 = Font(name="Times New Roman", size=10, bold=False)

    worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)
    worksheet.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=max_col)
    worksheet.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=max_col)

    worksheet.cell(row=title_row, column=1).alignment = center_align
    worksheet.cell(row=title_row, column=1).font = font_12_bold

    for col in range(1, max_col + 1):
        worksheet.cell(row=title_row, column=col).border = title_border

    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.alignment = right_align
        cell.font = font_12_bold
        cell.border = item_header_border if col == 1 else header_border

    for row in range(body_start_row, body_end_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = right_align
            cell.font = font_12_bold if col == 1 else font_12

    note_text = worksheet.cell(row=notes_row, column=1).value
    worksheet.cell(row=notes_row, column=1).alignment = left_align
    if isinstance(note_text, str) and note_text.startswith("Note(s):"):
        note_body = note_text[len("Note(s):") :].lstrip()
        worksheet.cell(row=notes_row, column=1).value = CellRichText(
            TextBlock(InlineFont(rFont="Times New Roman", sz=10, b=True), "Note(s): "),
            TextBlock(InlineFont(rFont="Times New Roman", sz=10), note_body),
        )
    else:
        worksheet.cell(row=notes_row, column=1).font = font_10

    for col in range(1, max_col + 1):
        worksheet.cell(row=notes_row, column=col).border = notes_border

    source_text = worksheet.cell(row=source_row, column=1).value
    worksheet.cell(row=source_row, column=1).alignment = left_align
    if isinstance(source_text, str) and source_text.startswith("Source(s):"):
        source_body = source_text[len("Source(s):") :].lstrip()
        worksheet.cell(row=source_row, column=1).value = CellRichText(
            TextBlock(InlineFont(rFont="Times New Roman", sz=10, b=True), "Source(s): "),
            TextBlock(InlineFont(rFont="Times New Roman", sz=10), source_body),
        )
    else:
        worksheet.cell(row=source_row, column=1).font = font_10

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save(filename)


def presentable_table_to_excel_bytes(final_df: pd.DataFrame) -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        export_presentable_table_to_excel(final_df=final_df, filename=str(temp_path))
        return temp_path.read_bytes()
    finally:
        temp_path.unlink(missing_ok=True)


def export_selected_presentable_tables_to_excel(
    report: dict,
    metrics: Sequence[str],
    output_folder: str = "excel_tables",
) -> dict[str, str]:
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    saved_files = {}
    for metric in clean_metrics(metrics):
        final_df = report["presentable_tables"][metric]
        file_path = output_path / f"{metric.lower()}_table.xlsx"
        export_presentable_table_to_excel(final_df=final_df, filename=str(file_path))
        saved_files[metric] = str(file_path)

    return saved_files


def wrap_lines(text: str, width: int = 110) -> list[str]:
    lines = textwrap.wrap(
        str(text),
        width=width,
        break_long_words=False,
        break_on_hyphens=False,
    )
    return lines if lines else [""]


def get_preferred_chart_font_family() -> str:
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    for candidate in CHART_FONT_CANDIDATES:
        if candidate in available_fonts:
            return candidate
    return "serif"


def build_figure_texts(
    report: dict,
    metric: str,
    source_text: str = "Compustat - North America",
) -> tuple[str, str, str]:
    long_table = report["long_table"].copy()
    company_names = long_table.loc[long_table["tic"] != "", "conm"].drop_duplicates().tolist()
    company_text = join_company_names(company_names)

    years = sorted(report["pivot_tables"][metric].columns.tolist())
    first_year = years[0]
    last_year = years[-1]
    metric_label = get_metric_label(metric)

    title = f"{metric_label} for {company_text} in {first_year}-{last_year}"
    notes = (
        f"Note(s): This figure presents the {metric_label} for {company_text} "
        f"in the years of {first_year} to {last_year}."
    )
    source = f"Source(s): {source_text}"
    return title, notes, source


def export_metric_chart_to_svg(
    report: dict,
    metric: str,
    filename: str | None = None,
    source_text: str = "Compustat - North America",
    percent_metrics: Sequence[str] | None = None,
    figsize: tuple[int, int] = (10, 6),
):
    metric = metric.upper()
    percent_metrics = [item.upper() for item in (percent_metrics or DEFAULT_PERCENT_METRICS)]

    if metric not in report["pivot_tables"]:
        raise ValueError(f"{metric} was not found in the report.")

    pivot_df = report["pivot_tables"][metric].copy()
    title, notes, source = build_figure_texts(report, metric, source_text=source_text)

    note_label = "Note(s):"
    note_body = notes.replace(note_label, "").strip()
    source_label = "Source(s):"
    source_body = source.replace(source_label, "").strip()

    note_lines = wrap_lines(note_body, width=105)
    source_lines = wrap_lines(source_body, width=105)
    chart_font_family = get_preferred_chart_font_family()

    fig, ax = plt.subplots(figsize=figsize)

    for row_name in pivot_df.index:
        series = pivot_df.loc[row_name]
        x_values = pivot_df.columns
        if "SIC" in str(row_name).upper() and "AVERAGE" in str(row_name).upper():
            ax.plot(x_values, series, marker="o", linestyle="--", linewidth=2, label=row_name)
        else:
            ax.plot(x_values, series, marker="o", linewidth=2.2, label=row_name)

    ax.set_title(
        title,
        fontsize=12,
        fontweight="bold",
        fontfamily=chart_font_family,
        pad=12,
    )
    ax.set_xlabel("Fiscal Year", fontsize=10, fontfamily=chart_font_family)
    ax.set_ylabel(get_metric_label(metric), fontsize=10, fontfamily=chart_font_family)
    ax.set_xticks(list(pivot_df.columns))
    ax.set_xticklabels([str(year) for year in pivot_df.columns], fontsize=10, fontfamily=chart_font_family)

    for label in ax.get_yticklabels():
        label.set_fontsize(10)
        label.set_fontfamily(chart_font_family)

    if metric in percent_metrics:
        ax.yaxis.set_major_formatter(PercentFormatter(1.0))

    legend = ax.legend(prop={"family": chart_font_family, "size": 10})
    if legend is not None:
        for text in legend.get_texts():
            text.set_fontfamily(chart_font_family)
            text.set_fontsize(10)

    ax.grid(True, linestyle=":", linewidth=0.8)

    line_gap = 0.028
    section_gap = 0.02
    footer_bottom = 0.02
    note_block_lines = max(len(note_lines), 1)
    source_block_lines = max(len(source_lines), 1)
    source_top_y = footer_bottom + (source_block_lines - 1) * line_gap
    note_top_y = source_top_y + section_gap + note_block_lines * line_gap
    body_x = 0.10

    for index, line in enumerate(note_lines):
        y_pos = note_top_y - index * line_gap
        if index == 0:
            fig.text(
                0.01,
                y_pos,
                note_label,
                ha="left",
                va="bottom",
                fontsize=10,
                fontweight="bold",
                fontfamily=chart_font_family,
            )
        fig.text(
            body_x,
            y_pos,
            line,
            ha="left",
            va="bottom",
            fontsize=10,
            fontfamily=chart_font_family,
        )

    for index, line in enumerate(source_lines):
        y_pos = source_top_y - index * line_gap
        if index == 0:
            fig.text(
                0.01,
                y_pos,
                source_label,
                ha="left",
                va="bottom",
                fontsize=10,
                fontweight="bold",
                fontfamily=chart_font_family,
            )
        fig.text(
            body_x + 0.02,
            y_pos,
            line,
            ha="left",
            va="bottom",
            fontsize=10,
            fontfamily=chart_font_family,
        )

    footer_height = note_top_y + 0.035
    bottom_margin = max(0.16, footer_height)
    plt.tight_layout(rect=[0, bottom_margin, 1, 1])

    if filename is not None:
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(filename, format="svg", bbox_inches="tight")

    return fig, ax


def metric_chart_to_svg_bytes(
    report: dict,
    metric: str,
    source_text: str = "Compustat - North America",
    percent_metrics: Sequence[str] | None = None,
    figsize: tuple[int, int] = (10, 6),
) -> bytes:
    fig, _ = export_metric_chart_to_svg(
        report=report,
        metric=metric,
        source_text=source_text,
        percent_metrics=percent_metrics,
        figsize=figsize,
    )
    try:
        buffer = io.BytesIO()
        fig.savefig(buffer, format="svg", bbox_inches="tight")
        return buffer.getvalue()
    finally:
        plt.close(fig)


def export_selected_metric_charts_to_svg(
    report: dict,
    metrics: Sequence[str],
    output_folder: str = "svg_charts",
    source_text: str = "Compustat - North America",
    percent_metrics: Sequence[str] | None = None,
    figsize: tuple[int, int] = (10, 6),
) -> dict[str, str]:
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    saved_files = {}
    for metric in clean_metrics(metrics):
        file_path = output_path / f"{metric.lower()}_chart.svg"
        fig, _ = export_metric_chart_to_svg(
            report=report,
            metric=metric,
            filename=str(file_path),
            source_text=source_text,
            percent_metrics=percent_metrics,
            figsize=figsize,
        )
        plt.close(fig)
        saved_files[metric] = str(file_path)

    return saved_files


def create_openai_compatible_client(base_url: str, api_key: str, use_system_proxy: bool = False) -> OpenAI:
    base_url = base_url.strip().rstrip("/")
    api_key = api_key.strip()
    if not base_url:
        raise ValueError("Please provide a base URL for the OpenAI-compatible API.")
    if not api_key:
        raise ValueError("Please provide an API key for the OpenAI-compatible API.")
    # Default to a direct connection, but allow users to opt into system proxy settings when needed.
    http_client = httpx.Client(trust_env=use_system_proxy, timeout=120.0)
    return OpenAI(base_url=base_url, api_key=api_key, http_client=http_client)


def create_lmstudio_client(base_url: str = "http://localhost:1234/v1", api_key: str = "lm-studio") -> OpenAI:
    return create_openai_compatible_client(base_url=base_url, api_key=api_key)


def get_available_lmstudio_models(client: OpenAI) -> list[str]:
    try:
        return [model.id for model in client.models.list().data]
    except Exception as exc:
        raise RuntimeError(
            "Could not connect to LM Studio. Please make sure LM Studio is open, "
            "a model is loaded, and the local server is running."
        ) from exc


def resolve_model_id(client: OpenAI, preferred_model: str = "gemma-4-e4b-it") -> str:
    available_models = get_available_lmstudio_models(client)
    if not available_models:
        raise RuntimeError("No model is currently available in LM Studio. Please load a local model first.")
    if preferred_model in available_models:
        return preferred_model
    return available_models[0]


def normalize_metric_list(report: dict, metrics: Sequence[str] | str) -> list[str]:
    available_metrics = list(report["markdown_tables"].keys())

    if metrics is None:
        raise ValueError("Please provide a metric list.")

    if isinstance(metrics, str):
        metrics = [metrics]

    metrics_upper = [metric.upper() for metric in metrics]
    if "ALL" in metrics_upper:
        return available_metrics

    missing = [metric for metric in metrics_upper if metric not in available_metrics]
    if missing:
        raise ValueError(
            "These metrics are not available in report['markdown_tables']: "
            f"{missing}. Available metrics are: {available_metrics}"
        )

    return metrics_upper


def build_combined_tables_block(report: dict, metrics: Sequence[str] | str) -> str:
    metrics = normalize_metric_list(report, metrics)
    blocks = []

    for index, metric in enumerate(metrics, start=1):
        metric_label = get_metric_label(metric)
        title_text = report["titles"].get(metric, metric_label)
        note_text = report["notes"].get(metric, "")
        source_text = report["sources"].get(metric, "")
        markdown_table = report["markdown_tables"][metric]

        block = f"""
==============================
Table {index}: {metric_label}
==============================
Title: {title_text}
{note_text}
{source_text}

Markdown table:
{markdown_table}
""".strip()
        blocks.append(block)

    return "\n\n".join(blocks)


def build_integrated_multi_table_prompt(
    report: dict,
    metrics: Sequence[str] | str,
    analysis_style: str = "entry-level financial analyst",
) -> str:
    metrics = normalize_metric_list(report, metrics)
    combined_tables_text = build_combined_tables_block(report, metrics)
    metric_labels = [get_metric_label(metric) for metric in metrics]
    metric_text = ", ".join(metric_labels)

    long_df = report["long_table"].copy()
    company_names = long_df.loc[long_df["tic"] != "", "conm"].drop_duplicates().tolist()
    company_text = join_company_names(company_names)
    company_count = len(company_names)
    benchmark_rows = long_df.loc[long_df["tic"] == "", "conm"].drop_duplicates().tolist()
    has_benchmark = bool(benchmark_rows)
    benchmark_text = join_company_names(benchmark_rows) if benchmark_rows else "No benchmark"

    all_years = set()
    for metric in metrics:
        all_years.update(report["pivot_tables"][metric].columns.tolist())

    year_list = sorted(all_years)
    first_year = year_list[0]
    last_year = year_list[-1]

    if company_count <= 1 and has_benchmark:
        table_scope_text = (
            "You are given related financial tables for one company together with an industry-average benchmark."
        )
        task_text = (
            "Your job is to produce one integrated analysis section that focuses on the company first "
            "and uses the benchmark only as supporting context."
        )
        instruction_lines = [
            "- Focus on the single company rather than writing as if multiple companies were compared.",
            "- Use the industry-average benchmark as context when it is informative.",
            "- Comment on gaps versus the benchmark only when the tables clearly support that comparison.",
        ]
    elif company_count <= 1:
        table_scope_text = "You are given related financial tables for one company."
        task_text = (
            "Your job is to produce one integrated analysis section about that company across the requested metrics."
        )
        instruction_lines = [
            "- Focus on the single company rather than writing as if multiple companies were compared.",
            "- Discuss changes across time and across metrics for that company when the evidence supports it.",
            "- Do not refer to peer comparisons or benchmark comparisons unless they are explicitly shown.",
        ]
    elif has_benchmark:
        table_scope_text = (
            "You are given related financial tables for multiple companies together with an industry-average benchmark."
        )
        task_text = (
            "Your job is to produce one integrated analysis section that compares the companies with each other "
            "and with the benchmark when the evidence supports it."
        )
        instruction_lines = [
            "- Compare the companies with each other across the requested metrics when relevant.",
            "- Use the industry-average benchmark as additional context where it helps interpretation.",
            "- Do not force a benchmark comparison in every point if the tables do not support it.",
        ]
    else:
        table_scope_text = "You are given related financial tables for multiple companies."
        task_text = (
            "Your job is to produce one integrated analysis section that compares the companies across the requested metrics."
        )
        instruction_lines = [
            "- Compare the companies with each other across the requested metrics when relevant.",
            "- Use cross-metric patterns and time trends to make the comparison more meaningful.",
            "- Do not describe missing benchmark evidence because no benchmark is shown.",
        ]

    scope_lines = "\n".join(instruction_lines)

    return f"""
Act as a careful {analysis_style}.

{table_scope_text}
{task_text}

Requested metrics:
{metric_text}

Coverage period:
{first_year} to {last_year}

Companies:
{company_text}

Industry benchmark:
{benchmark_text}

Important instructions:
- Use only the information visible in the tables below.
- Analyse the tables together, not one by one in isolation.
- Compare patterns across the requested metrics when relevant.
- Do not invent company events, external causes, management decisions, or numbers not shown.
- If the evidence is limited, say so clearly.
- Keep the writing professional, specific, and concise.
- Prefer financially meaningful interpretation over generic statements.
{scope_lines}

Output format:
### Integrated Analysis

#### 3 Observations Across the Tables
1.
2.
3.

#### 2 Risks Across the Tables
1.
2.

#### 1 Deeper Analysis Direction
1.

Tables:
{combined_tables_text}
""".strip()


def _extract_chat_completion_text(response) -> str:
    message = response.choices[0].message
    content = getattr(message, "content", "")

    if isinstance(content, str):
        return content

    if isinstance(content, list):
        text_parts = []
        for item in content:
            if isinstance(item, dict):
                if item.get("type") == "text" and item.get("text"):
                    text_parts.append(str(item["text"]))
            else:
                text_value = getattr(item, "text", None)
                if text_value:
                    text_parts.append(str(text_value))
        return "\n".join(text_parts).strip()

    return str(content)


def _extract_response_api_text(response) -> str:
    output_text = getattr(response, "output_text", None)
    if output_text:
        return str(output_text)

    output_items = getattr(response, "output", None) or []
    text_parts = []
    for item in output_items:
        item_type = getattr(item, "type", None)
        if item_type == "message":
            for content_item in getattr(item, "content", []) or []:
                content_type = getattr(content_item, "type", None)
                if content_type in {"output_text", "text"}:
                    text_value = getattr(content_item, "text", None)
                    if text_value:
                        text_parts.append(str(text_value))
    return "\n".join(text_parts).strip()


def _build_analysis_system_instruction() -> str:
    return (
        "You are a careful financial analyst. "
        "Only use the tables and instructions provided by the user. "
        "Do not make up unsupported facts."
    )


def ask_openai_compatible_for_integrated_analysis(
    report: dict,
    metrics: Sequence[str] | str,
    model_id: str,
    base_url: str,
    api_key: str,
    temperature: float = 0.2,
    max_tokens: int = 1400,
    analysis_style: str = "entry-level financial analyst",
    enable_thinking: bool = False,
    use_system_proxy: bool = False,
    provider_name: str = "OpenAI-compatible API",
    api_style: str = "chat_completions",
) -> dict:
    metrics = normalize_metric_list(report, metrics)

    client = create_openai_compatible_client(
        base_url=base_url,
        api_key=api_key,
        use_system_proxy=use_system_proxy,
    )
    prompt = build_integrated_multi_table_prompt(
        report=report,
        metrics=metrics,
        analysis_style=analysis_style,
    )

    if api_style == "responses":
        request_kwargs = {
            "model": model_id,
            "instructions": _build_analysis_system_instruction(),
            "input": prompt,
            "temperature": temperature,
            "max_output_tokens": max_tokens,
        }
        if enable_thinking:
            request_kwargs["extra_body"] = {"enable_thinking": True}

        response = client.responses.create(**request_kwargs)
        analysis_text = _extract_response_api_text(response)
    else:
        request_kwargs = {
            "model": model_id,
            "messages": [
                {
                    "role": "system",
                    "content": _build_analysis_system_instruction(),
                },
                {
                    "role": "user",
                    "content": prompt,
                },
            ],
            "temperature": temperature,
            "max_tokens": max_tokens,
        }
        if enable_thinking:
            request_kwargs["extra_body"] = {"enable_thinking": True}

        response = client.chat.completions.create(**request_kwargs)
        analysis_text = _extract_chat_completion_text(response)

    return {
        "metrics": metrics,
        "model_id": model_id,
        "provider_name": provider_name,
        "prompt": prompt,
        "analysis": analysis_text,
        "combined_tables_text": build_combined_tables_block(report, metrics),
        "enable_thinking": enable_thinking,
        "use_system_proxy": use_system_proxy,
        "api_style": api_style,
    }


def ask_lmstudio_for_integrated_analysis(
    report: dict,
    metrics: Sequence[str] | str,
    preferred_model: str = "gemma-4-e4b-it",
    base_url: str = "http://localhost:1234/v1",
    api_key: str = "lm-studio",
    temperature: float = 0.2,
    max_tokens: int = 1400,
    analysis_style: str = "entry-level financial analyst",
) -> dict:
    metrics = normalize_metric_list(report, metrics)

    client = create_lmstudio_client(base_url=base_url, api_key=api_key)
    model_id = resolve_model_id(client, preferred_model=preferred_model)
    return ask_openai_compatible_for_integrated_analysis(
        report=report,
        metrics=metrics,
        model_id=model_id,
        base_url=base_url,
        api_key=api_key,
        temperature=temperature,
        max_tokens=max_tokens,
        analysis_style=analysis_style,
        enable_thinking=False,
        provider_name="LM Studio",
    )


def save_integrated_analysis_result(result: dict, output_folder: str = "llm_outputs") -> dict[str, str]:
    output_path = Path(output_folder)
    output_path.mkdir(parents=True, exist_ok=True)

    safe_name = "_".join(metric.lower() for metric in result["metrics"][:5])
    if len(result["metrics"]) > 5:
        safe_name += "_etc"

    prompt_file = output_path / f"{safe_name}_integrated_prompt.md"
    tables_file = output_path / f"{safe_name}_combined_tables.md"
    analysis_file = output_path / f"{safe_name}_integrated_analysis.md"
    meta_file = output_path / f"{safe_name}_integrated_meta.json"

    prompt_file.write_text(result["prompt"], encoding="utf-8")
    tables_file.write_text(result["combined_tables_text"], encoding="utf-8")
    analysis_file.write_text(result["analysis"], encoding="utf-8")
    meta_file.write_text(
        json.dumps({"metrics": result["metrics"], "model_id": result["model_id"]}, indent=2),
        encoding="utf-8",
    )

    return {
        "prompt_file": str(prompt_file),
        "tables_file": str(tables_file),
        "analysis_file": str(analysis_file),
        "meta_file": str(meta_file),
    }


def analyse_requested_tables_with_lmstudio(
    report: dict,
    metrics: Sequence[str] | str,
    preferred_model: str = "gemma-4-e4b-it",
    base_url: str = "http://localhost:1234/v1",
    api_key: str = "lm-studio",
    temperature: float = 0.2,
    max_tokens: int = 1400,
    analysis_style: str = "entry-level financial analyst",
    output_folder: str = "llm_outputs",
    save_result: bool = True,
) -> dict:
    result = ask_lmstudio_for_integrated_analysis(
        report=report,
        metrics=metrics,
        preferred_model=preferred_model,
        base_url=base_url,
        api_key=api_key,
        temperature=temperature,
        max_tokens=max_tokens,
        analysis_style=analysis_style,
    )

    if save_result:
        result["saved_files"] = save_integrated_analysis_result(result=result, output_folder=output_folder)

    return result


def analyse_requested_tables_with_openai_compatible(
    report: dict,
    metrics: Sequence[str] | str,
    model_id: str,
    base_url: str,
    api_key: str,
    temperature: float = 0.2,
    max_tokens: int = 1400,
    analysis_style: str = "entry-level financial analyst",
    output_folder: str = "llm_outputs",
    save_result: bool = True,
    enable_thinking: bool = False,
    use_system_proxy: bool = False,
    provider_name: str = "OpenAI-compatible API",
    api_style: str = "chat_completions",
) -> dict:
    result = ask_openai_compatible_for_integrated_analysis(
        report=report,
        metrics=metrics,
        model_id=model_id,
        base_url=base_url,
        api_key=api_key,
        temperature=temperature,
        max_tokens=max_tokens,
        analysis_style=analysis_style,
        enable_thinking=enable_thinking,
        use_system_proxy=use_system_proxy,
        provider_name=provider_name,
        api_style=api_style,
    )

    if save_result:
        result["saved_files"] = save_integrated_analysis_result(result=result, output_folder=output_folder)

    return result


def build_analysis_markdown_bundle(report: dict, metrics: Sequence[str] | str) -> str:
    metrics = normalize_metric_list(report, metrics)
    parts = []
    for metric in metrics:
        parts.append(f"## {report['titles'][metric]}")
        parts.append(report["notes"][metric])
        parts.append(report["sources"][metric])
        parts.append("")
        parts.append(report["markdown_tables"][metric])
        parts.append("")
    return "\n".join(parts).strip()


def get_metric_reference_rows() -> list[dict[str, str]]:
    rows = []
    for group_name, metrics in METRIC_GROUPS.items():
        for metric in metrics:
            reference = METRIC_REFERENCE.get(metric, {})
            rows.append(
                {
                    "group": group_name,
                    "code": metric,
                    "metric": get_metric_label(metric),
                    "formula": reference.get("formula", ""),
                    "meaning": reference.get("meaning", ""),
                    "notes": reference.get("notes", ""),
                }
            )
    return rows


def list_metric_labels(metrics: Iterable[str]) -> list[str]:
    return [get_metric_label(metric) for metric in metrics]
